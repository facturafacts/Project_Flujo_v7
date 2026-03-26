"""
Excel Export — v7 Multi-Account Labeling with Weekly P&L

Sheets:
  1. 📊 Summary       — stats by account
  2. 🔵 Account A      — unlabeled from A  ← LABEL HERE
  3. 🟢 Account B      — unlabeled from B  ← LABEL HERE
  4. ✅ All Labeled    — review / un-label
  5. 📂 Catalog       — Chart of Accounts (edit col C → P&L formulas update)
  6. 💰 P&L           — weekly, SUMIFS from Account A + Account B sheets
  [hidden] Dropdowns  — source data for dropdown menus (do not edit)

P&L Design:
  - P&L SUMIFS read directly from Account A + Account B sheets
    (NOT a separate combined sheet — no external link warnings)
  - Category names live in Catalog col C only
  - P&L section rows list catalog row numbers (→ row 8) for reference
  - To rename a category: change ONE cell in Catalog col C → done

Usage:
    python scripts/excel/export.py
"""
import sys, os, datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from data.db_manager import (
    get_connection, get_unlabeled_for_export, get_labeled_for_review, init_db,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────
THIN   = Side(style="thin",   color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size)
def aln(h="left", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

HDR_FILL    = fill("343A40")
BLUE_A_FILL = fill("CCE5FF")
GREEN_B_FILL= fill("D4EDDA")
YELLOW_FILL = fill("FFF3CD")
INCOME_FILL = fill("D1ECF1")
EXPENSE_FILL= fill("FCE4D6")
WHITE_FILL  = fill("FFFFFF")

def hdr_cell(cell, text=""):
    cell.value = text; cell.fill = HDR_FILL
    cell.font  = font(bold=True, color="FFFFFF")
    cell.alignment = aln("center"); cell.border = BORDER

def sc(cell, value=None, fill_=None, bold=False, color="000000",
        h="left", size=10, fmt=None, wrap=False):
    if value is not None: cell.value = value
    if fill_:             cell.fill = fill_
    cell.font      = font(bold=bold, color=color, size=size)
    cell.alignment = aln(h=h, wrap=wrap)
    cell.border    = BORDER
    if fmt: cell.number_format = fmt

def set_cols(ws, widths):
    for col, w in widths.items(): ws.column_dimensions[col].width = w


# ════════════════════════════════════════════════════════════════
# CATALOG — single source of truth for all subcategory names
# ════════════════════════════════════════════════════════════════
CATALOG = [
    # (Context,    Direction,  Subcategory)
    ("Context",    "Direction", "Subcategory"),          # row 2 = header
    ("Work",       "Inflow",   "Client Payment"),
    ("Work",       "Inflow",   "Retainer"),
    ("Work",       "Inflow",   "Capital Injection"),
    ("Work",       "Inflow",   "Reimbursement"),
    ("Work",       "Inflow",   "Interest / Yield"),
    ("Work",       "Outflow",  "Liquor"),
    ("Work",       "Outflow",  "Wine Supplier"),
    ("Work",       "Outflow",  "Craft Beer"),
    ("Work",       "Outflow",  "Butcher"),
    ("Work",       "Outflow",  "Groceries"),
    ("Work",       "Outflow",  "Ginger Beer"),
    ("Work",       "Outflow",  "Septic Maint"),
    ("Work",       "Outflow",  "Payroll"),
    ("Work",       "Outflow",  "Kitchen / Cutlery"),
    ("Work",       "Outflow",  "Software / Subscriptions"),
    ("Work",       "Outflow",  "Contractors / Freelancers"),
    ("Work",       "Outflow",  "Equipment"),
    ("Work",       "Outflow",  "Marketing / Ads"),
    ("Work",       "Outflow",  "Banking Fees"),
    ("Work",       "Outflow",  "Utilities"),
    ("Work",       "Outflow",  "Rent"),
    ("Work",       "Outflow",  "Professional Services"),
    ("Work",       "Outflow",  "Taxes & Licenses"),
    ("Work",       "Outflow",  "Travel"),
    ("Work",       "Outflow",  "Fuel"),
    ("Work",       "Outflow",  "Maintenance"),
    ("Work",       "Outflow",  "Insurance"),
    ("Work",       "Outflow",  "Other"),
    ("Personal",   "Inflow",   "Wallet Top-Up"),
    ("Personal",   "Inflow",   "Owner Draw"),
    ("Personal",   "Inflow",   "Refund / Reimbursement"),
    ("Personal",   "Inflow",   "Gift / Repayment"),
    ("Personal",   "Outflow",  "Groceries"),
    ("Personal",   "Outflow",  "Restaurants"),
    ("Personal",   "Outflow",  "Personal Shopping"),
    ("Personal",   "Outflow",  "Health / Wellness"),
    ("Personal",   "Outflow",  "Travel"),
    ("Personal",   "Outflow",  "Entertainment"),
    ("Personal",   "Outflow",  "Education"),
    ("Personal",   "Outflow",  "Other"),
]
CATALOG_ROW = {name: idx + 2 for idx, (_, _, name) in enumerate(CATALOG) if name != "Subcategory"}


# ════════════════════════════════════════════════════════════════
# SHEET 1 — SUMMARY
# ════════════════════════════════════════════════════════════════
def write_summary(ws, total, unlabeled, work, personal, by_acc):
    ws.sheet_view.showGridLines = False
    hdr_cell(ws["A1"], "Mercado Pago — Labeling Status  (v7)")
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws.merge_cells("A1:D1"); ws.row_dimensions[1].height = 30

    for c, h in enumerate(["", "Account A (Expense)", "Account B (Concentrator)", "Total"], 1):
        hdr_cell(ws.cell(3, c), h)

    rows_ = [
        ("Total entries",     by_acc.get("A", 0), by_acc.get("B", 0), total),
        ("Unlabeled",         "—",               "—",                unlabeled),
        ("Labeled Work",      "—",               "—",                work),
        ("Labeled Personal",  "—",               "—",                personal),
    ]
    for r, (lbl, a, b, tot) in enumerate(rows_, 4):
        sc(ws.cell(r, 1), lbl, bold=True)
        for c, v in enumerate([a, b, tot], 2):
            sc(ws.cell(r, c), v, h="center")
        for c in range(1, 5): ws.cell(r, c).border = BORDER

    sc(ws.cell(8, 1), "Instructions", bold=True, size=12)
    lines = [
        "1. Open Account A or Account B sheet to label transactions.",
        "2. Col D = Classification (Work / Personal) ← dropdown",
        "3. Col E = Subcategory ← dropdown (adapts to Classification)",
        "4. Yellow rows = intercompany (auto-detected, skip these).",
        "5. When done: save file → run: python scripts/excel/import.py",
        "6. Weekly P&L is on the 💰 P&L sheet — live, updates when you press F9.",
    ]
    for i, line in enumerate(lines, 9):
        ws.merge_cells(f"A{i}:D{i}")
        sc(ws.cell(i, 1), line, color="555555", size=9)
    set_cols(ws, {"A": 30, "B": 22, "C": 26, "D": 14})


# ════════════════════════════════════════════════════════════════
# SHEET 2 & 3 — TO LABEL (Account A and B) + DROPDOWNS
# ════════════════════════════════════════════════════════════════
LABEL_HDRS = [
    "internal_id", "Date", "Category", "Classification",
    "Subcategory", "Description", "Gross", "Fee", "Net", "Intercompany"
]
# Hidden dropdown data (written to hidden "Dropdowns" sheet, referenced by DataValidation)
WORK_SUBS = [n for (_, d, n) in CATALOG if d == "Outflow" and n != "Subcategory"]
PERS_SUBS = [n for (_, d, n) in CATALOG if d == "Inflow" and "Personal" not in CATALOG[CATALOG.index((_, d, n)) - 1][0] and n not in ("Wallet Top-Up", "Owner Draw", "Refund / Reimbursement", "Gift / Repayment")]
# Actually compute correctly:
PERS_SUBS = [n for (_, d, n) in CATALOG if d == "Inflow" and n not in ("Client Payment", "Retainer", "Capital Injection", "Reimbursement", "Interest / Yield")]
PERS_SUBS_OUT = [n for (_, d, n) in CATALOG if d == "Outflow" and n not in WORK_SUBS]
ALL_SUBS = [n for (_, _, n) in CATALOG if n not in ("Subcategory", "Context", "Direction")]


def build_dropdowns_sheet(ws):
    """Hidden sheet — source data for all dropdown menus."""
    ws.sheet_state = "hidden"
    ws["A1"] = "Classification"; ws["A1"].font = Font(bold=True)
    for i, v in enumerate(["Work", "Personal"], 2):
        ws.cell(i, 1, v)

    ws["C1"] = "Work_Subcategory"; ws["C1"].font = Font(bold=True)
    for i, v in enumerate(WORK_SUBS, 2):
        ws.cell(i, 3, v)

    ws["E1"] = "Personal_Subcategory"; ws["E1"].font = Font(bold=True)
    for i, v in enumerate(PERS_SUBS + PERS_SUBS_OUT, 2):
        ws.cell(i, 5, v)


def write_label_sheet(ws, account, rows, wb):
    ws.sheet_view.showGridLines = False
    acct_fill = BLUE_A_FILL if account == "A" else GREEN_B_FILL
    tag = "Account A — To Label" if account == "A" else "Account B — To Label"
    icon = "🔵" if account == "A" else "🟢"

    hdr_cell(ws["A1"], f"{icon}  {tag}  — fill Classification + Subcategory columns")
    ws["A1"].font = Font(bold=True, size=12, color="FFFFFF")
    ws.merge_cells("A1:I1"); ws.row_dimensions[1].height = 26

    for c, h in enumerate(LABEL_HDRS, 1):
        hdr_cell(ws.cell(3, c), h)

    last_data_row = len(rows) + 3

    for r_idx, row in enumerate(rows, 4):
        iid, sa, date, cat, desc, gross, fee, net, interco = row
        row_fill = YELLOW_FILL if interco else acct_fill
        vals = [
            iid,
            (str(date)[:10] if date else ""),
            cat or "",
            "",    # col D — Classification
            "",    # col E — Subcategory
            desc or "",
            gross or 0,
            fee or 0,
            (gross or 0) + (fee or 0),  # col I — Net = gross + fee
            "⚠️ INTERCO" if interco else ""
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r_idx, c, v)
            cell.fill = row_fill; cell.border = BORDER
            if c in (7, 8, 9) and isinstance(v, (int, float)):
                cell.number_format = '"$"#,##0.00'
            if c == 10 and interco:
                cell.alignment = aln("center")

    # ── Data Validations (dropdowns) ──────────────────────────
    # Only add dropdowns if there are rows to label
    if last_data_row > 3:
        dv_cls = DataValidation(
            type="list",
            formula1="Dropdowns!$A$2:$A$3",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid",
            error="Choose Work or Personal",
            showInputMessage=True,
            promptTitle="Classification",
            prompt="Choose: Work or Personal"
        )
        dv_cls.add(f"D4:D{last_data_row}")
        ws.add_data_validation(dv_cls)

        dv_sub = DataValidation(
            type="list",
            formula1="Dropdowns!$C$2:$C$32",
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Invalid subcategory",
            error="Choose a subcategory from the list"
        )
        dv_sub.add(f"E4:E{last_data_row}")
        ws.add_data_validation(dv_sub)

    set_cols(ws, {"A": 36, "B": 12, "C": 22, "D": 16, "E": 28,
                  "F": 38, "G": 12, "H": 14, "I": 12, "J": 14, "K": 14})


# ════════════════════════════════════════════════════════════════
# SHEET 4 — ALL LABELED (Review)
# ════════════════════════════════════════════════════════════════
def write_review_sheet(ws, rows):
    ws.sheet_view.showGridLines = False
    hdr_cell(ws["A1"], "✅  All Labeled Transactions (read-only)")
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws.merge_cells("A1:K1"); ws.row_dimensions[1].height = 26

    for c, h in enumerate(["internal_id","Acct","Date","Category","Classification",
                             "Subcategory","Description","Gross","Fee","Net","Interco"], 1):
        hdr_cell(ws.cell(3, c), h)

    for r_idx, row in enumerate(rows, 4):
        iid, sa, date, cat, subcat, cls, desc, gross, fee, net, interco = row
        rfill = YELLOW_FILL if interco else (BLUE_A_FILL if sa == "A" else GREEN_B_FILL)
        for c, v in enumerate([iid, sa, (str(date)[:10] if date else ""), cat,
                                 cls, subcat, desc, gross, fee, net,
                                 "⚠️" if interco else ""], 1):
            cell = ws.cell(r_idx, c, v)
            cell.fill = rfill; cell.border = BORDER
            if c in (8, 9, 10) and isinstance(v, (int, float)):
                cell.number_format = '"$"#,##0.00'

    set_cols(ws, {"A": 36, "B": 6, "C": 12, "D": 22, "E": 14, "F": 28,
                  "G": 38, "H": 14, "I": 12, "J": 14, "K": 8})


# ════════════════════════════════════════════════════════════════
# SHEET 5 — CATALOG
# ════════════════════════════════════════════════════════════════
def write_catalog_sheet(ws):
    ws.sheet_view.showGridLines = False
    hdr_cell(ws["A1"],
             "📂  Chart of Accounts — SOURCE OF TRUTH  |  "
             "Edit col C to rename a category → all P&L formulas update automatically")
    ws["A1"].font = Font(bold=True, size=11, color="FFFFFF")
    ws.merge_cells("A1:C1"); ws.row_dimensions[1].height = 28

    for c, h in enumerate(["Context", "Direction", "Subcategory  ← edit here to rename"], 1):
        hdr_cell(ws.cell(2, c), h)

    ctx_fills = {"Work": fill("D1ECF1"), "Personal": fill("F8D7DA"), "Context": fill("E9ECEF")}

    for r_idx, (ctx, dr, name) in enumerate(CATALOG, 3):
        f = ctx_fills.get(ctx, WHITE_FILL)
        for c, v in enumerate([ctx, dr, name], 1):
            cell = ws.cell(r_idx, c, v)
            cell.fill = f; cell.border = BORDER; cell.alignment = aln()
        if r_idx == 3:
            ws.cell(r_idx, 1).font = Font(italic=True, color="888888")

    note_row = len(CATALOG) + 4
    ws.merge_cells(f"A{note_row}:C{note_row}")
    sc(ws.cell(note_row, 1),
       "💡 To rename a category: change the name in col C above. "
       "No formula editing needed — P&L SUMIFS reference these cells.",
       color="555555", size=9, wrap=True)
    ws.row_dimensions[note_row].height = 28
    set_cols(ws, {"A": 14, "B": 12, "C": 38})


# ════════════════════════════════════════════════════════════════
# SHEET 6 — P&L WEEKLY
# P&L reads DIRECTLY from Account A + Account B sheets (no separate combined sheet)
# ════════════════════════════════════════════════════════════════
MONEY = '"$"#,##0.00'

def to_week_key(dt) -> str:
    if not dt: return ""
    s = str(dt).strip()[:10]
    try:
        from pandas import Timestamp
        ts = Timestamp(s)
        if str(ts) == "NaT": return ""
        y, wk, _ = ts.isocalendar()
        return f"{y}-W{wk:02d}"
    except Exception:
        return ""


def sumifs_cat(src_sheet, sum_col, cls, cat_ref, week_col_l, sum_col_l):
    """
    SUMIFS over src_sheet for a given classification + subcategory from Catalog.
    src_sheet: "Account A" or "Account B"
    cat_ref:   e.g. "Catalog!$C7"
    week_col_l: e.g. "G" (the week key column in the label sheets)
    sum_col_l:  the column letter to sum
    Returns the SUMIFS formula string.
    """
    if week_col_l:
        return (f"=IFERROR(SUMIFS('{src_sheet}'!${sum_col_l}:${sum_col_l},"
                f"'{src_sheet}'!$D:$D,\"Work\","
                f"'{src_sheet}'!$E:$E,{cat_ref},"
                f"'{src_sheet}'!$G:$G,{week_col_l}$2),0)")
    else:  # ALL TIME — no week filter
        return (f"=IFERROR(SUMIFS('{src_sheet}'!${sum_col_l}:${sum_col_l},"
                f"'{src_sheet}'!$D:$D,\"Work\","
                f"'{src_sheet}'!$E:$E,{cat_ref}),0)")


def sumifs_cat_personal(src_sheet, sum_col_l, cat_ref, week_col_l):
    if week_col_l:
        return (f"=IFERROR(SUMIFS('{src_sheet}'!${sum_col_l}:${sum_col_l},"
                f"'{src_sheet}'!$D:$D,\"Personal\","
                f"'{src_sheet}'!$E:$E,{cat_ref},"
                f"'{src_sheet}'!$G:$G,{week_col_l}$2),0)")
    else:
        return (f"=IFERROR(SUMIFS('{src_sheet}'!${sum_col_l}:${sum_col_l},"
                f"'{src_sheet}'!$D:$D,\"Personal\","
                f"'{src_sheet}'!$E:$E,{cat_ref}),0)")


def build_pnl_sheet(ws, df_all):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B4"

    # ── Collect week keys from all data ──────────────────────
    df_dated = df_all[df_all["date"].apply(to_week_key) != ""].copy()
    df_dated["week_key"] = df_dated["date"].apply(to_week_key)
    all_weeks = sorted(df_dated["week_key"].unique())
    recent_weeks = all_weeks[-13:] if len(all_weeks) > 13 else all_weeks

    DATA_COL_START = 3   # col C = first week
    LAST_WEEK_COL  = DATA_COL_START + len(recent_weeks) - 1
    TOTAL_COL      = LAST_WEEK_COL + 1
    TOTAL_COL_L    = get_column_letter(TOTAL_COL)

    # ── Row 1: title ─────────────────────────────────────────
    ws.merge_cells(f"A1:{TOTAL_COL_L}1")
    ws["A1"].value = "P62  WEEKLY PROFIT & LOSS"
    ws["A1"].font  = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill  = fill("1F3864")
    ws["A1"].alignment = aln("center")
    ws.row_dimensions[1].height = 28

    # ── Row 2: week headers ─────────────────────────────────
    hdr_cell(ws.cell(2, 1), "LINE ITEM")
    for wi, wk in enumerate(recent_weeks, DATA_COL_START):
        hdr_cell(ws.cell(2, wi), wk)
    hdr_cell(ws.cell(2, TOTAL_COL), "ALL TIME")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 2
    for ci in range(DATA_COL_START, TOTAL_COL + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 11

    def section_bar(row, label):
        ws.cell(row, 1, f"  {label}")
        for ci in range(1, TOTAL_COL + 1):
            c = ws.cell(row=row, column=ci)
            c.fill = fill("D9E1F2"); c.border = BORDER
            c.font = Font(bold=True, color="1F3864", size=10)
            c.alignment = aln("left")

    def write_data_row(ws, row, label, formula_a, formula_b, font_color, bg):
        """Write a row that SUMs across both Account A and Account B sheets."""
        c = ws.cell(row, 1, f"  {label}")
        c.font = Font(size=10, color=font_color); c.fill = fill(bg)
        c.border = BORDER; c.alignment = aln("left")

        for ci in range(DATA_COL_START, TOTAL_COL + 1):
            cl = get_column_letter(ci)
            is_total = (ci == TOTAL_COL)
            f_a = formula_a(None if is_total else cl)
            f_b = formula_b(None if is_total else cl)
            formula = f"={f_a[1:]}+{f_b[1:]}" if f_a.startswith("=") else f_a
            # Combine: add both sheets
            combined = f"=IFERROR({f_a[1:]}+{f_b[1:]},0)"
            cell = ws.cell(row, ci, combined)
            cell.number_format = MONEY
            cell.font = Font(size=10, color=font_color)
            cell.fill = fill(bg); cell.border = BORDER
            cell.alignment = aln(h="right")

    def total_row_fn(ws, row, label, component_rows, font_color, bg):
        c = ws.cell(row, 1, f"  {label}")
        c.font = Font(bold=True, size=10, color=font_color)
        c.fill = fill(bg); c.border = BORDER; c.alignment = aln("left")
        for ci in range(DATA_COL_START, TOTAL_COL + 1):
            cl = get_column_letter(ci)
            refs = "+".join([f"{cl}{r}" for r in component_rows])
            cell = ws.cell(row, ci, f"={refs}")
            cell.number_format = MONEY
            cell.font = Font(bold=True, size=10, color=font_color)
            cell.fill = fill(bg); cell.border = BORDER
            cell.alignment = aln(h="right")
        return row

    def write_cat_row(ws, row, name, font_color, bg):
        """One expense/income row from catalog, summed across A + B."""
        cr = CATALOG_ROW.get(name)
        if not cr: return None
        cat_ref = f"Catalog!$C{cr}"
        label = f"  {name}"

        c = ws.cell(row, 1, label)
        c.font = Font(size=10, color=font_color)
        c.fill = fill(bg); c.border = BORDER; c.alignment = aln("left")

        for ci in range(DATA_COL_START, TOTAL_COL + 1):
            cl = get_column_letter(ci)
            is_total = (ci == TOTAL_COL)
            if is_total:
                f_a = sumifs_cat("Account A", "K", "Work", cat_ref, None, "K")
                f_b = sumifs_cat("Account B", "K", "Work", cat_ref, None, "K")
            else:
                f_a = sumifs_cat("Account A", "K", "Work", cat_ref, cl, "K")
                f_b = sumifs_cat("Account B", "K", "Work", cat_ref, cl, "K")
            combined = f"=IFERROR({f_a[1:]}+{f_b[1:]},0)"
            cell = ws.cell(row, ci, combined)
            cell.number_format = MONEY
            cell.font = Font(size=10, color=font_color)
            cell.fill = fill(bg); cell.border = BORDER
            cell.alignment = aln(h="right")
        return row

    row = 3

    # ── WORK INCOME ──────────────────────────────────────────
    section_bar(row, "WORK  INCOME"); row += 1
    income_rows = []

    for name in ["Client Payment", "Retainer", "Interest / Yield"]:
        cr = CATALOG_ROW.get(name)
        if not cr: continue
        cat_ref = f"Catalog!$C{cr}"
        c = ws.cell(row, 1, f"  {name}")
        c.font = Font(size=10, color="0070C0"); c.fill = fill("DEEAF1")
        c.border = BORDER; c.alignment = aln("left")
        for ci in range(DATA_COL_START, TOTAL_COL + 1):
            cl = get_column_letter(ci)
            is_total = (ci == TOTAL_COL)
            if is_total:
                f_a = sumifs_cat("Account A", "K", "Work", cat_ref, None, "K")
                f_b = sumifs_cat("Account B", "K", "Work", cat_ref, None, "K")
            else:
                f_a = sumifs_cat("Account A", "K", "Work", cat_ref, cl, "K")
                f_b = sumifs_cat("Account B", "K", "Work", cat_ref, cl, "K")
            cell = ws.cell(row, ci, f"=IFERROR({f_a[1:]}+{f_b[1:]},0)")
            cell.number_format = MONEY; cell.font = Font(size=10, color="0070C0")
            cell.fill = fill("DEEAF1"); cell.border = BORDER; cell.alignment = aln(h="right")
        income_rows.append(row); row += 1

    # POS Sales — category = "POS Sale" in col C of label sheets
    c = ws.cell(row, 1, "  POS Sales"); c.font = Font(size=10, color="0070C0")
    c.fill = fill("DEEAF1"); c.border = BORDER; c.alignment = aln("left")
    for ci in range(DATA_COL_START, TOTAL_COL + 1):
        cl = get_column_letter(ci)
        is_total = (ci == TOTAL_COL)
        if is_total:
            f = (f"=IFERROR(SUMIFS('Account A'!$K:$K,'Account A'!$D:$D,\"Work\","
                 f"'Account A'!$C:$C,\"POS Sale\",'Account A'!$K:$K,\">0\")"
                 f"+SUMIFS('Account B'!$K:$K,'Account B'!$D:$D,\"Work\","
                 f"'Account B'!$C:$C,\"POS Sale\",'Account B'!$K:$K,\">0\"),0)")
        else:
            f = (f"=IFERROR(SUMIFS('Account A'!$K:$K,'Account A'!$D:$D,\"Work\","
                 f"'Account A'!$C:$C,\"POS Sale\",'Account A'!$G:$G,{cl}$2,'Account A'!$K:$K,\">0\")"
                 f"+SUMIFS('Account B'!$K:$K,'Account B'!$D:$D,\"Work\","
                 f"'Account B'!$C:$C,\"POS Sale\",'Account B'!$G:$G,{cl}$2,'Account B'!$K:$K,\">0\"),0)")
        cell = ws.cell(row, ci, f)
        cell.number_format = MONEY; cell.font = Font(size=10, color="0070C0")
        cell.fill = fill("DEEAF1"); cell.border = BORDER; cell.alignment = aln(h="right")
    income_rows.append(row); row += 1

    row = total_row_fn(ws, row, "Work Income Total", income_rows, "0070C0", "BDD7EE"); row += 2

    # ── WORK EXPENSES ─────────────────────────────────────────
    expense_rows = []
    section_bar(row, "WORK  EXPENSES"); row += 1

    for name in WORK_SUBS:
        r = write_cat_row(ws, row, name, "C00000", "FCE4D6")
        if r: expense_rows.append(r); row += 1

    # MP Fees
    c = ws.cell(row, 1, "  MP Fees (Work)"); c.font = Font(size=10, color="C00000")
    c.fill = fill("FCE4D6"); c.border = BORDER; c.alignment = aln("left")
    for ci in range(DATA_COL_START, TOTAL_COL + 1):
        cl = get_column_letter(ci)
        is_total = (ci == TOTAL_COL)
        if is_total:
            f = ("=IFERROR(SUMIFS('Account A'!$H:$H,'Account A'!$D:$D,\"Work\")"
                 "+SUMIFS('Account B'!$H:$H,'Account B'!$D:$D,\"Work\"),0)")
        else:
            f = (f"=IFERROR(SUMIFS('Account A'!$H:$H,'Account A'!$D:$D,\"Work\","
                 f"'Account A'!$G:$G,{cl}$2)"
                 f"+SUMIFS('Account B'!$H:$H,'Account B'!$D:$D,\"Work\","
                 f"'Account B'!$G:$G,{cl}$2),0)")
        cell = ws.cell(row, ci, f)
        cell.number_format = MONEY; cell.font = Font(size=10, color="C00000")
        cell.fill = fill("FCE4D6"); cell.border = BORDER; cell.alignment = aln(h="right")
    expense_rows.append(row); row += 1

    row = total_row_fn(ws, row, "Work Expenses Total", expense_rows, "C00000", "F4B8B8"); row += 2

    # ── NET PROFIT ───────────────────────────────────────────
    it_r = income_rows[-1]; et_r = expense_rows[-1]
    for ci in range(DATA_COL_START, TOTAL_COL + 1):
        cl = get_column_letter(ci)
        cell = ws.cell(row, ci, f"={cl}{it_r}+{cl}{et_r}")
        cell.number_format = MONEY; cell.fill = fill("1F3864")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.border = BORDER; cell.alignment = aln(h="right")
    ws.cell(row, 1, "  NET PROFIT").font = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row, 1).fill = fill("1F3864"); ws.cell(row, 1).border = BORDER
    net_row = row; row += 2

    # ── PERSONAL CASH FLOW ────────────────────────────────────
    section_bar(row, "PERSONAL  (Cash Flow — not P&L)"); row += 1
    personal_rows = []
    for name in ["Wallet Top-Up", "Owner Draw"] + PERS_SUBS_OUT:
        cr = CATALOG_ROW.get(name)
        if not cr: continue
        cat_ref = f"Catalog!$C{cr}"
        c = ws.cell(row, 1, f"  {name}")
        c.font = Font(size=10, color="7B7B7B"); c.fill = fill("F8F9FA")
        c.border = BORDER; c.alignment = aln("left")
        for ci in range(DATA_COL_START, TOTAL_COL + 1):
            cl = get_column_letter(ci)
            is_total = (ci == TOTAL_COL)
            if is_total:
                f_a = sumifs_cat_personal("Account A", "K", cat_ref, None)
                f_b = sumifs_cat_personal("Account B", "K", cat_ref, None)
            else:
                f_a = sumifs_cat_personal("Account A", "K", cat_ref, cl)
                f_b = sumifs_cat_personal("Account B", "K", cat_ref, cl)
            cell = ws.cell(row, ci, f"=IFERROR({f_a[1:]}+{f_b[1:]},0)")
            cell.number_format = MONEY; cell.font = Font(size=10, color="7B7B7B")
            cell.fill = fill("F8F9FA"); cell.border = BORDER; cell.alignment = aln(h="right")
        personal_rows.append(row); row += 1

    if personal_rows:
        row = total_row_fn(ws, row, "Personal Cash Flow", personal_rows, "7B7B7B", "E9ECEF"); row += 2

    # ── FOOTNOTES ─────────────────────────────────────────────
    w0 = recent_weeks[0] if recent_weeks else "—"
    w1 = recent_weeks[-1] if recent_weeks else "—"
    notes = [
        f"Notes:  • Last 13 weeks: {w0} → {w1}",
        "• Income (blue): Work + net > 0  |  Expenses (red): Work + net < 0  |  Gray: Personal",
        "• To rename a category: change the name in Catalog sheet col C — all formulas update automatically",
        "• P&L reads directly from Account A + Account B sheets — no external links",
        "• Label transactions → save → press F9 to refresh P&L",
    ]
    for ni, note_text in enumerate(notes, row):
        ws.merge_cells(f"A{ni}:{TOTAL_COL_L}{ni}")
        c = ws.cell(ni, 1, note_text)
        c.font = Font(italic=True, size=9, color="888888")
        c.alignment = Alignment(wrap_text=True, horizontal="left")
        ws.row_dimensions[ni].height = 16

    return recent_weeks


# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════
def run_export():
    init_db()
    conn = get_connection()

    unlabeled_a = get_unlabeled_for_export("A", limit=10000)
    unlabeled_b = get_unlabeled_for_export("B", limit=10000)

    cur = conn.cursor()
    cur.execute("""
        SELECT internal_id, source_account, date, category,
               subcategory, description, gross_amount, mp_fee, net_amount,
               classification, source
        FROM ledger_final ORDER BY date DESC
    """)
    all_rows = cur.fetchall()
    conn.close()

    import pandas as pd
    cols = ["internal_id","source_account","date","category",
            "subcategory","description","gross_amount","mp_fee","net_amount",
            "classification","source"]
    df_all = pd.DataFrame(all_rows, columns=cols)

    total    = len(df_all)
    unlabeled = int((df_all["classification"].isna() | (df_all["classification"] == "")).sum())
    work     = int((df_all["classification"] == "Work").sum())
    personal = int((df_all["classification"] == "Personal").sum())
    by_acc   = df_all.groupby("source_account").size().to_dict()
    labeled  = get_labeled_for_review()

    ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    out_path = Path(__file__).parent.parent.parent / "output" / f"v7_labeling_{ts}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # Build sheets
    ws1 = wb.create_sheet("📊 Summary")
    ws2 = wb.create_sheet("🔵 Account A")
    ws3 = wb.create_sheet("🟢 Account B")
    ws4 = wb.create_sheet("✅ All Labeled")
    ws5 = wb.create_sheet("📂 Catalog")
    ws6 = wb.create_sheet("💰 P&L")
    ws_drop = wb.create_sheet("Dropdowns")   # hidden

    build_dropdowns_sheet(ws_drop)
    write_summary(ws1, total, unlabeled, work, personal, by_acc)
    write_label_sheet(ws2, "A", unlabeled_a, wb)
    write_label_sheet(ws3, "B", unlabeled_b, wb)
    write_review_sheet(ws4, labeled)
    write_catalog_sheet(ws5)
    weeks = build_pnl_sheet(ws6, df_all)

    # Add week_key column (col G) to Account A + B sheets so SUMIFS can filter by week
    for ws_lbl in (ws2, ws3):
        hdr_cell(ws_lbl.cell(3, 7), "Week")
        for r in range(4, ws_lbl.max_row + 1):
            date_val = ws_lbl.cell(r, 2).value  # col B = date
            if date_val:
                ws_lbl.cell(r, 7, to_week_key(date_val))

    wb.save(out_path)
    print(f"\n✅  Exported: {out_path.name}")
    print(f"    Account A: {len(unlabeled_a)} unlabeled")
    print(f"    Account B: {len(unlabeled_b)} unlabeled")
    print(f"    Total labeled: {len(labeled)}")
    print(f"    P&L weeks: {weeks[0]} → {weeks[-1]}  ({len(weeks)} columns)")


if __name__ == "__main__":
    run_export()
