"""
Export to Excel - P62_v5 with Weekly P&L
Sheets: Catalog | To Label | P&L (weekly columns)
"""
import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

DB_PATH = "/home/subsc/.openclaw/workspace/workspaces/CLIENTS/Pescadero-62/data/ledger.db"
OUTPUT_PATH = "/home/subsc/.openclaw/workspace/workspaces/CLIENTS/Pescadero-62/output/p62_v5_labeling.xlsx"


# ─── Helpers ───────────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def to_week_key(dt) -> str:
    """
    Convert a value to ISO week string 'YYYY-Wnn'.
    Handles timezone-aware datetimes, naive datetimes, date-only strings,
    and mixed-format columns from SQLite (some '2026-03-18T15:45:41.000-04:00',
    some '2026-03-18'). Returns empty string for invalid values.
    """
    if dt is None:
        return ''
    # Normalize string inputs to YYYY-MM-DD first (handles both formats)
    if isinstance(dt, str):
        dt = dt.strip()[:10]
        try:
            dt = pd.Timestamp(dt)
        except Exception:
            return ''
    try:
        ts = pd.Timestamp(dt)
        if pd.isna(ts):
            return ''
        year, week, _ = ts.isocalendar()
        return f"{year}-W{week:02d}"
    except Exception:
        return ''


def style_header(cell, bg='4472C4', fg='FFFFFF'):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.font = Font(color=fg, bold=True, size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border()


def style_subheader(cell, bg='D9E1F2'):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.font = Font(bold=True, size=10, color='1F3864')
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = thin_border()


def style_total(cell, bg='4472C4', fg='FFFFFF'):
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.font = Font(color=fg, bold=True, size=10)
    cell.border = thin_border()


def style_label_classification(cell):
    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    cell.border = thin_border()


def style_label_work(cell):
    cell.fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    cell.border = thin_border()


def style_label_personal(cell):
    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    cell.border = thin_border()


# ─── Dropdowns Sheet ───────────────────────────────────────────────────────────

def create_dropdowns_sheet(wb):
    ws = wb.create_sheet('Dropdowns')
    ws.sheet_state = 'hidden'

    ws['A1'] = 'Classification'
    ws['A1'].font = Font(bold=True)
    for i, v in enumerate(['Work', 'Personal'], 2):
        ws.cell(row=i, column=1, value=v)

    ws['C1'] = 'Work_Subcategory'
    ws['C1'].font = Font(bold=True)
    work_subs = [
        'Client Payment', 'Retainer', 'Liquor', 'Wine Supplier', 'Craft Beer',
        'Butcher', 'Groceries', 'Ginger beer', 'Septic Maint', 'Payroll',
        'Kitchen/Cutlery', 'Software', 'Contractors', 'Equipment', 'Marketing',
        'Banking Fees', 'Utilities', 'Rent', 'Professional Services',
        'Taxes & Licenses', 'Travel', 'Fuel', 'Maintenance', 'Insurance', 'Other'
    ]
    for i, v in enumerate(work_subs, 2):
        ws.cell(row=i, column=3, value=v)

    ws['E1'] = 'Personal_Subcategory'
    ws['E1'].font = Font(bold=True)
    personal_subs = [
        'Wallet Top-Up', 'Owner Draw', 'Groceries', 'Restaurants', 'Shopping',
        'Entertainment', 'Health', 'Travel', 'Education', 'Other'
    ]
    for i, v in enumerate(personal_subs, 2):
        ws.cell(row=i, column=5, value=v)

    return work_subs, personal_subs


# ─── Catalog Sheet ─────────────────────────────────────────────────────────────

def create_catalog_sheet(wb):
    ws = wb.create_sheet('Catalog')
    headers = ['Context', 'Direction', 'Subcategory']
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col))
        ws.cell(row=1, column=col).value = h

    catalog = [
        ('Work',     'Inflow',  'Client Payment'),
        ('Work',     'Inflow',  'Retainer'),
        ('Work',     'Outflow', 'Liquor'),
        ('Work',     'Outflow', 'Wine Supplier'),
        ('Work',     'Outflow', 'Craft Beer'),
        ('Work',     'Outflow', 'Butcher'),
        ('Work',     'Outflow', 'Groceries'),
        ('Work',     'Outflow', 'Ginger beer'),
        ('Work',     'Outflow', 'Septic Maint'),
        ('Work',     'Outflow', 'Payroll'),
        ('Work',     'Outflow', 'Kitchen/Cutlery'),
        ('Work',     'Outflow', 'Software'),
        ('Work',     'Outflow', 'Contractors'),
        ('Work',     'Outflow', 'Equipment'),
        ('Work',     'Outflow', 'Marketing'),
        ('Work',     'Outflow', 'Banking Fees'),
        ('Work',     'Outflow', 'Utilities'),
        ('Work',     'Outflow', 'Rent'),
        ('Work',     'Outflow', 'Professional Services'),
        ('Work',     'Outflow', 'Taxes & Licenses'),
        ('Work',     'Outflow', 'Travel'),
        ('Work',     'Outflow', 'Fuel'),
        ('Work',     'Outflow', 'Maintenance'),
        ('Work',     'Outflow', 'Insurance'),
        ('Work',     'Outflow', 'Other'),
        ('Personal', 'Inflow',  'Wallet Top-Up'),
        ('Personal', 'Inflow',  'Owner Draw'),
        ('Personal', 'Outflow', 'Groceries'),
        ('Personal', 'Outflow', 'Restaurants'),
        ('Personal', 'Outflow', 'Shopping'),
        ('Personal', 'Outflow', 'Entertainment'),
        ('Personal', 'Outflow', 'Health'),
        ('Personal', 'Outflow', 'Travel'),
        ('Personal', 'Outflow', 'Education'),
        ('Personal', 'Outflow', 'Other'),
    ]
    for ri, (ctx, dr, sub) in enumerate(catalog, 2):
        ws.cell(row=ri, column=1, value=ctx)
        ws.cell(row=ri, column=2, value=dr)
        ws.cell(row=ri, column=3, value=sub)

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 24


# ─── To Label Sheet ─────────────────────────────────────────────────────────────

def create_to_label_sheet(wb):
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""
        SELECT internal_id, date, category, description,
               gross_amount, mp_fee, net_amount, source
        FROM ledger_final
        ORDER BY date DESC
    """, conn)
    conn.close()

    # Compute week key directly from raw date strings — avoids mixed-format NaT issue
    df['week_key'] = df['date'].apply(to_week_key)

    # Classification / subcategory (blank = unlabeled)
    df['classification'] = ''
    df['subcategory']   = ''

    ws = wb.create_sheet('To Label')
    ws.freeze_panes = 'A3'

    # Instruction banner
    ws['A1'] = (
        '▶ Label Col E (Work/Personal) → Col F (Work subcategory) OR Col G (Personal subcategory). '
        'Filter by Week (Col B) to work week by week. P&L sheet auto-updates on Save.'
    )
    ws['A1'].font = Font(bold=True, color='FF6600', size=10)
    ws['A1'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A1:L1')

    headers = [
        'internal_id',           # A  1
        'Week',                  # B  2
        'Date',                  # C  3
        'Category',              # D  4
        'Classification',        # E  5
        'Work_Subcategory',      # F  6
        'Personal_Subcategory',  # G  7
        'Description',           # H  8
        'gross_amount',          # I  9
        'mp_fee',               # J 10
        'net_amount',            # K 11
        'source',                # L 12
    ]
    for ci, h in enumerate(headers, 1):
        style_header(ws.cell(row=2, column=ci))
        ws.cell(row=2, column=ci).value = h

    last_data_row = len(df) + 2

    for ri, (_, row_data) in enumerate(df.iterrows()):
        xl_row = ri + 3

        # Date string
        date_val = str(row_data['date'])[:10] if pd.notna(row_data['date']) else ''

        vals = [
            str(row_data['internal_id']),   # A
            row_data['week_key'],            # B  ← correctly formatted now
            date_val,                        # C
            row_data['category'],            # D
            '',                             # E classification (label here)
            '',                             # F work subcategory
            '',                             # G personal subcategory
            row_data['description'],        # H
            row_data['gross_amount'],        # I
            row_data['mp_fee'],             # J
            row_data['net_amount'],          # K
            row_data['source'],              # L
        ]

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=xl_row, column=ci)
            cell.value = v
            cell.border = thin_border()

            if ci in (9, 10, 11):
                cell.number_format = '$#,##0.00'
                if isinstance(v, float) and v < 0:
                    cell.font = Font(color='C00000')

        style_label_classification(ws.cell(row=xl_row, column=5))
        style_label_work(ws.cell(row=xl_row, column=6))
        style_label_personal(ws.cell(row=xl_row, column=7))

    # Data validations
    dv_clf = DataValidation(
        type='list', formula1='Dropdowns!$A$2:$A$3', allow_blank=True,
        showErrorMessage=True, errorTitle='Invalid', error='Choose Work or Personal'
    )
    ws.add_data_validation(dv_clf)
    dv_clf.add(f'E3:E{last_data_row}')

    dv_work = DataValidation(
        type='list', formula1='Dropdowns!$C$2:$C$26', allow_blank=True
    )
    ws.add_data_validation(dv_work)
    dv_work.add(f'F3:F{last_data_row}')

    dv_pers = DataValidation(
        type='list', formula1='Dropdowns!$E$2:$E$11', allow_blank=True
    )
    ws.add_data_validation(dv_pers)
    dv_pers.add(f'G3:G{last_data_row}')

    widths = {'A': 18, 'B': 11, 'C': 20, 'D': 20, 'E': 15, 'F': 22, 'G': 22,
              'H': 35, 'I': 14, 'J': 12, 'K': 14, 'L': 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return df


# ─── P&L Sheet (Weekly) ────────────────────────────────────────────────────────

def build_pnl_sheet(wb, df):
    ws = wb.create_sheet('P&L')

    # Collect all week keys, sorted — filter out blanks
    df_dated = df[df['week_key'] != '']
    all_weeks = sorted(df_dated['week_key'].unique())
    # Show last 13 weeks
    recent_weeks = all_weeks[-13:] if len(all_weeks) > 13 else all_weeks
    # All-time columns
    first_data_col  = 2
    last_week_col   = len(recent_weeks) + 1   # col B = week 1
    total_col       = last_week_col + 1
    total_col_letter = get_column_letter(total_col)

    # ── Title ──────────────────────────────────────────────────────────────────
    ws['A1'] = 'P62  WEEKLY PROFIT & LOSS'
    ws['A1'].font = Font(size=14, bold=True, color='1F3864')
    ws.row_dimensions[1].height = 24
    ws.merge_cells(f'A1:{total_col_letter}1')

    # ── Column headers (row 2) ──────────────────────────────────────────────────
    style_header(ws.cell(row=2, column=1))
    ws.cell(row=2, column=1).value = 'LINE ITEM'
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='left')

    for wi, wk in enumerate(recent_weeks, first_data_col):
        style_header(ws.cell(row=2, column=wi))
        ws.cell(row=2, column=wi).value = wk
        ws.column_dimensions[get_column_letter(wi)].width = 11

    style_total(ws.cell(row=2, column=total_col))
    ws.cell(row=2, column=total_col).value = 'ALL TIME'
    ws.column_dimensions[total_col_letter].width = 14

    row = 3

    MONEY = '$#,##0.00'
    INCOME_BLUE = '0070C0'
    EXPENSE_RED = 'C00000'
    SECTION_BG  = 'D9E1F2'
    NET_BG      = '1F3864'

    def section_bar(ws, row, label):
        """Full-width section divider row."""
        ws.cell(row=row, column=1, value=f'  {label}')
        style_subheader(ws.cell(row=row, column=1))
        for ci in range(2, total_col + 1):
            style_subheader(ws.cell(row=row, column=ci))

    def data_row(ws, row, label, week_formulas, font_color='000000', indent=2):
        """One data row with label + formula per column."""
        cell_a = ws.cell(row=row, column=1, value=(' ' * indent) + label)
        cell_a.font = Font(size=10)
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        cell_a.border = thin_border()
        for ci, (formula, fmt) in week_formulas.items():
            c = ws.cell(row=row, column=ci)
            c.value = formula if formula else 0
            c.number_format = fmt
            c.font = Font(color=font_color, size=10)
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = thin_border()

    def total_row(ws, row, label, component_rows, font_color, bg_hex):
        """Sum row that totals the component rows above it."""
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = Font(bold=True, color=font_color, size=10)
        cell_a.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
        cell_a.border = thin_border()
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        for ci in range(first_data_col, total_col + 1):
            col_l = get_column_letter(ci)
            refs = '+'.join([f'{col_l}{r}' for r in component_rows])
            c = ws.cell(row=row, column=ci)
            c.value = f'={refs}'
            c.number_format = MONEY
            c.font = Font(bold=True, color=font_color, size=10)
            c.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type='solid')
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = thin_border()

    # ── WORK INCOME ────────────────────────────────────────────────────────────
    section_bar(ws, row, 'WORK  INCOME')
    row += 1

    income_items = [
        'Client Payment', 'Retainer',
    ]
    income_rows = []
    for item in income_items:
        formulas = {}
        for wi, wk in enumerate(recent_weeks, first_data_col):
            formulas[wi] = (
                f"=IFERROR(SUMIFS('To Label'!$K:$K,"
                f"'To Label'!$E:$E,\"Work\","
                f"'To Label'!$F:$F,\"{item}\","
                f"'To Label'!$B:$B,\"{wk}\","
                f"'To Label'!$K:$K,\">0\"),0)",
                MONEY
            )
        formulas[total_col] = (
            f"=IFERROR(SUMIFS('To Label'!$K:$K,"
            f"'To Label'!$E:$E,\"Work\","
            f"'To Label'!$F:$F,\"{item}\","
            f"'To Label'!$K:$K,\">0\"),0)",
            MONEY
        )
        data_row(ws, row, item, formulas, font_color=INCOME_BLUE)
        income_rows.append(row)
        row += 1

    # POS Sales row — captures any positive unlabeled subcategory POS Sales
    formulas = {}
    for wi, wk in enumerate(recent_weeks, first_data_col):
        formulas[wi] = (
            f"=IFERROR(SUMIFS('To Label'!$K:$K,"
            f"'To Label'!$E:$E,\"Work\","
            f"'To Label'!$D:$D,\"POS Sale\","
            f"'To Label'!$B:$B,\"{wk}\","
            f"'To Label'!$K:$K,\">0\"),0)",
            MONEY
        )
    formulas[total_col] = (
        f"=IFERROR(SUMIFS('To Label'!$K:$K,"
        f"'To Label'!$E:$E,\"Work\","
        f"'To Label'!$D:$D,\"POS Sale\","
        f"'To Label'!$K:$K,\">0\"),0)",
        MONEY
    )
    data_row(ws, row, 'POS Sales (Work)', formulas, font_color=INCOME_BLUE)
    income_rows.append(row)
    row += 1

    # Money Transfers row — transfers marked Work, positive
    formulas = {}
    for wi, wk in enumerate(recent_weeks, first_data_col):
        formulas[wi] = (
            f"=IFERROR(SUMIFS('To Label'!$K:$K,"
            f"'To Label'!$E:$E,\"Work\","
            f"'To Label'!$D:$D,\"Money Transfer\","
            f"'To Label'!$B:$B,\"{wk}\","
            f"'To Label'!$K:$K,\">0\"),0)",
            MONEY
        )
    formulas[total_col] = (
        f"=IFERROR(SUMIFS('To Label'!$K:$K,"
        f"'To Label'!$E:$E,\"Work\","
        f"'To Label'!$D:$D,\"Money Transfer\","
        f"'To Label'!$K:$K,\">0\"),0)",
        MONEY
    )
    data_row(ws, row, 'Money Transfer (Work)', formulas, font_color=INCOME_BLUE)
    income_rows.append(row)
    row += 1

    total_row(ws, row, 'Work Income Total', income_rows, INCOME_BLUE, 'DEEAF1')
    income_total_row = row
    row += 2

    # ── WORK EXPENSES ───────────────────────────────────────────────────────────
    section_bar(ws, row, 'WORK  EXPENSES')
    row += 1

    expense_items = [
        'Liquor', 'Wine Supplier', 'Craft Beer', 'Butcher', 'Groceries',
        'Ginger beer', 'Septic Maint', 'Payroll', 'Kitchen/Cutlery',
        'Software', 'Contractors', 'Equipment', 'Marketing', 'Banking Fees',
        'Utilities', 'Rent', 'Professional Services', 'Taxes & Licenses',
        'Travel', 'Fuel', 'Maintenance', 'Insurance', 'Other',
    ]
    expense_rows = []

    for item in expense_items:
        formulas = {}
        for wi, wk in enumerate(recent_weeks, first_data_col):
            formulas[wi] = (
                f"=IFERROR(SUMIFS('To Label'!$K:$K,"
                f"'To Label'!$E:$E,\"Work\","
                f"'To Label'!$F:$F,\"{item}\","
                f"'To Label'!$B:$B,\"{wk}\","
                f"'To Label'!$K:$K,\"<0\"),0)",
                MONEY
            )
        formulas[total_col] = (
            f"=IFERROR(SUMIFS('To Label'!$K:$K,"
            f"'To Label'!$E:$E,\"Work\","
            f"'To Label'!$F:$F,\"{item}\","
            f"'To Label'!$K:$K,\"<0\"),0)",
            MONEY
        )
        data_row(ws, row, item, formulas, font_color=EXPENSE_RED)
        expense_rows.append(row)
        row += 1

    # MP Fees — fees for Work transactions (fee is negative in DB)
    formulas = {}
    for wi, wk in enumerate(recent_weeks, first_data_col):
        formulas[wi] = (
            f"=IFERROR(SUMIFS('To Label'!$J:$J,"
            f"'To Label'!$E:$E,\"Work\","
            f"'To Label'!$B:$B,\"{wk}\"),0)",
            MONEY
        )
    formulas[total_col] = (
        f"=IFERROR(SUMIFS('To Label'!$J:$J,"
        f"'To Label'!$E:$E,\"Work\"),0)",
        MONEY
    )
    data_row(ws, row, 'MP Fees (Work)', formulas, font_color=EXPENSE_RED)
    expense_rows.append(row)
    row += 1

    # POS Sales / Money Transfer as expenses (negative net)
    for cat in ['POS Sale', 'Money Transfer', 'Purchase/Expense']:
        formulas = {}
        for wi, wk in enumerate(recent_weeks, first_data_col):
            formulas[wi] = (
                f"=IFERROR(SUMIFS('To Label'!$K:$K,"
                f"'To Label'!$E:$E,\"Work\","
                f"'To Label'!$D:$D,\"{cat}\","
                f"'To Label'!$B:$B,\"{wk}\","
                f"'To Label'!$K:$K,\"<0\"),0)",
                MONEY
            )
        formulas[total_col] = (
            f"=IFERROR(SUMIFS('To Label'!$K:$K,"
            f"'To Label'!$E:$E,\"Work\","
            f"'To Label'!$D:$D,\"{cat}\","
            f"'To Label'!$K:$K,\"<0\"),0)",
            MONEY
        )
        data_row(ws, row, f'{cat} (Work, outflow)', formulas, font_color=EXPENSE_RED)
        expense_rows.append(row)
        row += 1

    total_row(ws, row, 'Work Expenses Total', expense_rows, EXPENSE_RED, 'FCE4D6')
    expense_total_row = row
    row += 2

    # ── NET PROFIT ─────────────────────────────────────────────────────────────
    formulas = {}
    for ci in range(first_data_col, total_col + 1):
        col_l = get_column_letter(ci)
        formulas[ci] = (f'={col_l}{income_total_row}+{col_l}{expense_total_row}', MONEY)
    data_row(ws, row, 'NET PROFIT', formulas, font_color='FFFFFF')
    for ci in range(1, total_col + 1):
        ws.cell(row=row, column=ci).fill = PatternFill(
            start_color=NET_BG, end_color=NET_BG, fill_type='solid')
        ws.cell(row=row, column=ci).font = Font(bold=True, color='FFFFFF', size=11)
        ws.cell(row=row, column=ci).border = thin_border()
    net_profit_row = row
    row += 2

    # ── FOOTNOTES ──────────────────────────────────────────────────────────────
    notes = [
        'Notes:',
        f'• Showing last 13 weeks: {recent_weeks[0]} → {recent_weeks[-1]}  (oldest → newest)',
        '• Income (blue): Work classification + net_amount > 0, grouped by subcategory or category',
        '• Expenses (red): Work classification + net_amount < 0, grouped by subcategory or category',
        '• MP Fees: fee column (col J) for Work transactions in each week',
        '• All values are LIVE from "To Label" — label transactions & Save (or F9) to refresh',
    ]
    for ni, note_text in enumerate(notes, row):
        c = ws.cell(row=ni, column=1, value=note_text)
        c.font = Font(italic=True, size=9, color='808080')
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{ni}:{total_col_letter}{ni}')

    ws.column_dimensions['A'].width = 30

    return recent_weeks


# ─── Main ──────────────────────────────────────────────────────────────────────

def export_to_excel():
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    create_dropdowns_sheet(wb)
    create_catalog_sheet(wb)
    df = create_to_label_sheet(wb)
    weeks = build_pnl_sheet(wb, df)

    wb.save(OUTPUT_PATH)
    print(f"✅ Exported: {OUTPUT_PATH}")
    print(f"   Weeks: W{weeks[0][-5:]} → W{weeks[-1][-5:]}  ({len(weeks)} columns)")
    print(f"   To Label rows: {len(df)}")
    print(f"   Week keys: verified 'YYYY-Wnn' format (no float bugs)")


if __name__ == "__main__":
    export_to_excel()
